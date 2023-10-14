import openpyxl
import pandas as pd



def column_number_to_letter(column_number):
    dividend = column_number
    column_letter = ''
    while dividend > 0:
        modulo = (dividend - 1) % 26
        column_letter = chr(65 + modulo) + column_letter
        dividend = (dividend - modulo) // 26
    return column_letter



# Load the workbook
workbook = openpyxl.load_workbook('Eco.xlsx')

# Select a specific sheet (e.g., the first sheet)
sheet = workbook.active

#columns = [chr(i) for i in range(ord('F'), ord('BR') + 1)]

#data = {}  # Create an empty dictionary to store the data

# Iterate through the columns and extract data
#for column in columns:
#    values = []
#    header = sheet[f'{column}1'].value  # Use the first cell as the header
#    for row in range(2, sheet.max_row + 1):
#        cell = sheet[f'{column}{row}']
#        values.append(cell.value)
#    data[header] = values

search_number = str(int(input("Enter the number to search for: ")))
target_column = None

columns = [column_number_to_letter(i) for i in range(6, 191)]

for column in columns:
    first_cell_value = str(sheet[f'{column}1'].value)
    extracted_number = first_cell_value.split()[0]
    if extracted_number[1:] == search_number:
        target_column = column
        break

if target_column is not None:
    # Extract data from the column
    data = []
    for row in range(2, sheet.max_row + 1):
        cell = sheet[f'{target_column}{row}']
        data.append(cell.value)

    # Calculate the percentage
    total_cells = len(data)
    count_ones = data.count(1)
    percentage = (count_ones / total_cells) * 100

    print(f"Percentage of 1s in column {target_column}: {percentage:.2f}%")
else:
    print(f"Number {search_number} not found in the first cell of any column.")

# Close the workbook
workbook.close()
